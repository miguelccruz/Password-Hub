import os
from openpyxl import Workbook
import openpyxl
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QWidget,
    QLabel,
    QPushButton,
    QFormLayout,
    QLineEdit,
    QComboBox,
    QMessageBox,
    QStyle,
    QHBoxLayout
    )

from PyQt6.QtGui import QIcon

class ResultsWindow(QWidget):
    def __init__(self, search_text, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.setWindowTitle('Password Hub: Results')
        self.setWindowIcon(QIcon('cat.ico'))
        self.setGeometry(760, 440, 500, 200)

        results_layout = QFormLayout()
        self.setLayout(results_layout)

        # get values from excel file
        row_number, account_name_value, category_value, email_user_value, password_value, note_value= self.find_data(search_text)

        self.existing_name = account_name_value

        result_label = QLabel(f'Results for: {search_text}')
        result_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        results_layout.addWidget(result_label)

        # Category
        category_label = QLabel(f'Category: {category_value}')
        category_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        results_layout.addWidget(category_label)

        # Account Name
        account_name_label = QLabel(f'Account Name: {account_name_value}')
        account_name_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        results_layout.addWidget(account_name_label)

        # Email/Username
        email_user_label = QLabel(f'Email/Username: {email_user_value}')
        email_user_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        results_layout.addWidget(email_user_label)

        # Password
        password_label = QLabel(f'Password: {password_value}')
        password_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        results_layout.addWidget(password_label)

        # Note
        note_label = QLabel(f'Note: {note_value}')
        note_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        results_layout.addWidget(note_label)

        record_values = (
            category_value,
            account_name_value,
            email_user_value,
            password_value,
            note_value
        )

        update_button = QPushButton('Update Account')
        update_button.clicked.connect(lambda: self.open_update_window(row_number, record_values, self.existing_name))
        results_layout.addWidget(update_button)

        self.show()

    def find_data(self, search_text):
        file_path = 'accounts_data/accounts_data.xlsx'

        if not os.path.exists(file_path):
            workbook = Workbook()
            workbook.save(file_path)

        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_col=2, max_col=2, min_row=1):
            for cell in row:
                if search_text.lower() in str(cell.value).lower():
                    category_value = ws.cell(row=cell.row, column=1).value
                    email_user_value = ws.cell(row=cell.row, column=3).value
                    password_value = ws.cell(row=cell.row, column=4).value
                    note_value = ws.cell(row=cell.row, column=5).value
                    return cell.row, cell.value, category_value, email_user_value, password_value, note_value
                
        print('finding data')
        return None, None, None, None, None
    
    def open_update_window(self, row_number, record_values, existing_name):
        self.update_window = UpdateWindow(row_number=row_number, record_values=record_values, existing_name=existing_name)
        self.update_window.show()
        self.close()

class UpdateWindow(QWidget):
    def __init__(self, row_number, record_values, existing_name, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.setWindowTitle('Password Hub: Update Record')
        self.setGeometry(760, 440, 500, 300)

        form_layout = QFormLayout()
        self.setLayout(form_layout)

        # Unpack record values
        category_value, account_name_value, email_user_value, password_value, note_value = record_values

        # ComboBox
        self.create_set_dropdown(category_value)
        self.account_name = QLineEdit(account_name_value)
        self.email_user = QLineEdit(email_user_value)


        self.password = QLineEdit(password_value, echoMode=QLineEdit.EchoMode.Password)
        self.show_password_icon = QPushButton(self)
        pixmapi = QStyle.StandardPixmap.SP_VistaShield
        icon = self.style().standardIcon(pixmapi)
        self.show_password_icon.setIcon(QIcon(icon))
        self.show_password_icon.setCheckable(True)
        self.show_password_icon.clicked.connect(self.toggle_password_visibility)

        password_layout = QHBoxLayout()
        password_layout.addWidget(self.password)
        password_layout.addWidget(self.show_password_icon)

        self.note = QLineEdit(note_value)
        submit_button = QPushButton('Submit')
        submit_button.clicked.connect(lambda: self.update_row(row_number, existing_name))

        form_layout.addRow('Category:', self.dropdown_menu)
        form_layout.addRow('Account Name:', self.account_name)
        form_layout.addRow('Email/Username:', self.email_user)
        form_layout.addRow('Password:', password_layout)
        form_layout.addRow('Note:', self.note)
        form_layout.addRow(submit_button)

        self.show()

    def create_set_dropdown(self, category_value):

        self.dropdown_menu = QComboBox(self)
        self.dropdown_menu.addItems([
            'Bank',
            'Browser',
            'E-mail',
            'Game',
            'Home',
            'Jobs',
            'Others',
            'School',
            'Social Media',
            'York'
        ])
        self.dropdown_menu.setCurrentText(category_value)

    def update_row(self, row_number, existing_name):
        new_data = [
            self.dropdown_menu.currentText(), 
            self.account_name.text(), 
            self.email_user.text(), 
            self.password.text(),
            self.note.text()
        ]
        
        file_path = 'accounts_data/accounts_data.xlsx'
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # data_validation flag
        found_clone = False

        # checks if account name already exists in database
        for row in ws.iter_rows(min_col=2, max_col=2, min_row=1):
            for cell in row:
                if self.account_name.text() == str(cell.value).lower():
                    QMessageBox.critical(
                        self,
                        'Error',
                        'Account Name already exists in database'
                    )
                    found_clone = True
                    break
            if found_clone:
                break
                    
        if not found_clone:
            for col_num, value in enumerate(new_data, start=1):
                ws.cell(row=row_number, column=col_num, value=value)

            wb.save(filename=file_path)
            print(f'Successfully updated row {row_number} in the database')
            self.close()

            self.replace_account_name(existing_name)

    def replace_account_name(self, existing_name):
        txt_file_path = 'accounts_data/account_names.txt'
        try:
            with open(txt_file_path, 'r') as file:
                lines = file.readlines()
                
            for i, line in enumerate(lines):
                if line.strip() == existing_name:
                    lines[i] = self.account_name.text() + '\n'

            
            with open(txt_file_path, 'w') as file:
                file.writelines(lines)

        except FileNotFoundError:
            print('txt file not found')

    def toggle_password_visibility(self, checked):
        if checked:
            self.password.setEchoMode(QLineEdit.EchoMode.Normal)
        else:
            self.password.setEchoMode(QLineEdit.EchoMode.Password)
            

        



